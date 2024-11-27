"""
DocuLens: Tool for Automated Policy Documents Analysis

This application processes PDF files to count occurrences of specific words defined in an Excel file.
Results are saved to an output Excel file with detailed counts and statistics.

Authors: Mahyar Sharafi-Laleh, Sarah Anne Ganter
GitHub: Moonyar
"""

# ---- Import Libraries ----
import sys
import os
import re
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QLabel,
    QProgressBar, QFileDialog, QMessageBox
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QFont, QIcon
from PyQt5.QtWidgets import QStyleFactory
import openpyxl
import fitz  # PyMuPDF for PDF text extraction

# ---- Utility Functions ----
def preprocess_text(text):
    """
    Preprocess text to remove hyphenation and ensure consistent word matching.
    :param text: Raw text from PDF
    :return: Preprocessed text
    """
    return re.sub(r'[‐‑‒–—⁃]\n', '', text)

# ---- Background Processing Thread ----
class WorkerThread(QThread):
    """
    WorkerThread for processing PDFs in the background without freezing the UI.
    """
    update_progress = pyqtSignal(int)  # Signal to update the progress bar
    completed = pyqtSignal(str)       # Signal to indicate task completion

    def __init__(self, excel_file, pdf_folder, output_file):
        """
        Initialize the WorkerThread with input and output paths.
        :param excel_file: Path to the Excel file containing words to search
        :param pdf_folder: Directory containing PDF files
        :param output_file: Path to save the results
        """
        super().__init__()
        self.excel_file = excel_file
        self.pdf_folder = pdf_folder
        self.output_file = output_file

    def run(self):
        """
        Core logic executed in the background.
        Processes the Excel file, reads PDFs, counts word occurrences, and saves results.
        """
        try:
            # Read words from Excel
            headers, data_rows = self.read_words_from_excel(self.excel_file)
            words_list = list(set([row[0].lower() for row in data_rows]))

            # List all PDF files in the folder
            pdf_files = [f for f in os.listdir(self.pdf_folder) if f.lower().endswith('.pdf')]
            counts = {word: {pdf: 0 for pdf in pdf_files} for word in words_list}
            total_words_in_all_documents = 0
            total_words_per_document = {}

            # Process each PDF file
            for pdf in pdf_files:
                pdf_path = os.path.join(self.pdf_folder, pdf)
                counts_in_pdf, total_words = self.count_words_in_pdf(pdf_path, words_list)
                total_words_in_all_documents += total_words
                total_words_per_document[pdf] = total_words

                # Update counts for each word
                for word in words_list:
                    counts[word][pdf] = counts_in_pdf[word]

                # Update progress
                self.update_progress.emit(int((list(pdf_files).index(pdf) + 1) / len(pdf_files) * 100))

            # Save results to an Excel file
            self.save_results(headers, data_rows, counts, pdf_files, total_words_per_document, total_words_in_all_documents)
            self.completed.emit(f"Processing complete. Results saved to {self.output_file}")
        except Exception as e:
            self.completed.emit(f"An error occurred: {e}")

    def read_words_from_excel(self, file_path):
        """
        Read words from the provided Excel file.
        :param file_path: Path to the Excel file
        :return: Tuple containing headers and data rows
        """
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        data_rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        return headers, data_rows

    def count_words_in_pdf(self, pdf_path, words_list):
        """
        Count occurrences of each word in a single PDF file.
        :param pdf_path: Path to the PDF file
        :param words_list: List of words to search for
        :return: Dictionary of word counts and total words
        """
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        text = preprocess_text(text.lower())
        counts = {word: 0 for word in words_list}
        total_words = 0
        for word in words_list:
            word_regex = r'\b{}\b'.format(re.escape(word))
            word_count = len(re.findall(word_regex, text, re.IGNORECASE))
            counts[word] = word_count
            total_words += word_count
        return counts, total_words

    def save_results(self, headers, data_rows, counts, pdf_files, total_words_per_document, total_words_in_all_documents):
        """
        Save the results to an Excel file.
        :param headers: Headers from the input Excel
        :param data_rows: Data rows from the input Excel
        :param counts: Word counts in each PDF
        :param pdf_files: List of processed PDF files
        :param total_words_per_document: Total words in each PDF
        :param total_words_in_all_documents: Total words across all PDFs
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        # Create extended headers
        extended_headers = headers + [f"{pdf} Count" for pdf in pdf_files] + \
                           [f"{pdf} Mean" for pdf in pdf_files] + ["Total", "Overall Mean"]
        ws.append(extended_headers)

        # Add word count data to rows
        all_rows_data = []
        for row in data_rows:
            word = row[0].lower()
            row_data = row[:]
            total_count = 0
            for pdf in pdf_files:
                count = counts[word][pdf]
                mean_per_document = count / total_words_per_document[pdf] if total_words_per_document[pdf] > 0 else 0
                row_data.extend([count, mean_per_document])
                total_count += count
            overall_mean = total_count / total_words_in_all_documents if total_words_in_all_documents > 0 else 0
            row_data.extend([total_count, overall_mean])
            all_rows_data.append(row_data)

        # Sort by total count in descending order
        total_index = extended_headers.index("Total")
        all_rows_data.sort(key=lambda x: x[total_index], reverse=True)

        # Write sorted data to Excel
        for row_data in all_rows_data:
            ws.append(row_data)

        wb.save(self.output_file)

# ---- Main Application UI ----
class App(QWidget):
    """
    Main GUI application for user interaction.
    Provides file selection, progress updates, and execution control.
    """
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        """
        Initialize the user interface elements.
        """
        self.setGeometry(300, 300, 600, 400)
        self.setWindowTitle('DocuLens')

        # Set the window icon
        self.setWindowIcon(QIcon('logo.png'))

        self.setStyleSheet("""
                    QWidget {
                        background-color: #2E2E2E;
                        font-family: Helvetica;
                        color: white;
                    }
                    QPushButton {
                        background-color: #0078D7;
                        color: white;
                        font-size: 14px;
                        border-radius: 8px;
                        padding: 10px;
                    }
                    QPushButton:hover {
                        background-color: #005bb5;
                    }
                    QPushButton:disabled {
                        background-color: #d3d3d3;
                        color: #a9a9a9;
                    }
                    QPushButton#startBtn {
                        background-color: #28a745;
                        color: white;
                    }
                    QPushButton#startBtn:hover {
                        background-color: #218838;
                    }
                    QProgressBar {
                        border: 2px solid #0078D7;
                        border-radius: 5px;
                        text-align: center;
                        color: white;
                    }
                    QProgressBar::chunk {
                        background-color: #0078D7;
                        width: 20px;
                    }
                    QLabel#fileLabel {
                        background-color: #555555;
                        border-radius: 5px;
                        padding: 10px;
                        margin: 5px 0;
                        font-size: 14px;
                        text-align: center;
                        color: white;
                    }
                    QLabel#selectedFileLabel {
                        background-color: #28a745;
                        border-radius: 5px;
                        padding: 10px;
                        margin: 5px 0;
                        font-size: 14px;
                        color: white;
                        text-align: center;
                    }
                """)

        layout = QVBoxLayout()

        self.label_excel = QLabel('No Excel file selected')
        self.label_excel.setObjectName("fileLabel")
        self.btn_select_excel = QPushButton('Select Excel File')
        self.btn_select_excel.clicked.connect(self.select_excel_file)
        layout.addWidget(self.btn_select_excel)
        layout.addWidget(self.label_excel, alignment=Qt.AlignCenter)

        self.label_pdf = QLabel('No PDF folder selected')
        self.label_pdf.setObjectName("fileLabel")
        self.btn_select_pdf = QPushButton('Select PDF Folder')
        self.btn_select_pdf.clicked.connect(self.select_pdf_folder)
        layout.addWidget(self.btn_select_pdf)
        layout.addWidget(self.label_pdf, alignment=Qt.AlignCenter)

        self.label_output = QLabel('No file selected for output')
        self.label_output.setObjectName("fileLabel")
        self.btn_select_output = QPushButton('Save Output As')
        self.btn_select_output.clicked.connect(self.select_output_file)
        layout.addWidget(self.btn_select_output)
        layout.addWidget(self.label_output, alignment=Qt.AlignCenter)

        self.progress_bar = QProgressBar(self)
        layout.addWidget(self.progress_bar)

        self.btn_start = QPushButton('Start Processing')
        self.btn_start.setObjectName("startBtn")
        self.btn_start.clicked.connect(self.start_processing)
        layout.addWidget(self.btn_start)

        self.setLayout(layout)
        self.show()

    def select_excel_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)",
                                                   options=options)
        if file_name:
            self.excel_file = file_name
            self.label_excel.setObjectName("selectedFileLabel")
            self.label_excel.setText(f'Selected: {os.path.basename(file_name)}')
            self.label_excel.setStyle(self.style())

    def select_pdf_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select PDF Folder")
        if folder:
            self.pdf_folder = folder
            self.label_pdf.setObjectName("selectedFileLabel")
            self.label_pdf.setText(f'Selected: {os.path.basename(folder)}')
            self.label_pdf.setStyle(self.style())

    def select_output_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Output As", "", "Excel Files (*.xlsx)", options=options)
        if file_name:
            self.output_file = file_name
            self.label_output.setObjectName("selectedFileLabel")
            self.label_output.setText(f'Selected: {os.path.basename(file_name)}')
            self.label_output.setStyle(self.style())

    def start_processing(self):
        if not hasattr(self, 'excel_file') or not hasattr(self, 'pdf_folder') or not hasattr(self, 'output_file'):
            QMessageBox.warning(self, 'Warning',
                                'Please select an Excel file, a PDF folder, and an output file location')
            return

        self.btn_start.setEnabled(False)
        self.btn_start.setStyleSheet("""
                    QPushButton {
                        background-color: #d3d3d3;
                        color: #a9a9a9;
                    }
                """)

        self.thread = WorkerThread(self.excel_file, self.pdf_folder, self.output_file)
        self.thread.update_progress.connect(self.progress_bar.setValue)
        self.thread.completed.connect(self.on_completed)
        self.thread.start()

    def on_completed(self, message):
        self.btn_start.setEnabled(True)
        self.btn_start.setStyleSheet("""
                    QPushButton#startBtn {
                        background-color: #28a745;
                        color: white;
                    }
                    QPushButton#startBtn:hover {
                        background-color: #218838;
                    }
                """)
        QMessageBox.information(self, 'Process Complete', message)

# ---- Application Entry Point ----
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))
    ex = App()
    sys.exit(app.exec_())

